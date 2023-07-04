import pandas as pd
from openpyxl import load_workbook

class ExcelExport:

    @classmethod
    def export_excel(cls, excel_name, columns_map, export_list):
        with pd.ExcelWriter(excel_name + ".xlsx") as writer:
            pf_failed = cls.format_excel(columns_map, export_list)
            pf_failed.to_excel(writer, sheet_name="Sheet1", index=False)
            writer.save()  # 保存表格

    @classmethod
    def format_excel(cls, columns_map, export_data: list):
        pf = pd.DataFrame(list(export_data))  # 将字典列表转换为DataFrame
        pf.rename(columns=columns_map, inplace=True)
        pf.fillna(" ", inplace=True)  # 替换空单元格
        return pf

class ExcelReader:

    @classmethod
    def read_excel(cls, excel_name, sheet):
        excel = load_workbook(excel_name + ".xlsx")
        table = excel.get_sheet_by_name(sheet)
        return table


if __name__ == "__main__":

    # 导出excel
    # columns_map = {
    #     "domain_name": "域名",
    #     "product_name": "产品",
    #     "apply_name": "申请人",
    #     "apply_time": "申请时间",
    #     "itsm_id": "itsm_id",
    #     "apply_reason": "备注",
    #     "tech_name": "开发负责人",
    #     "ops_name": "运维负责人",
    # }
    # record_list = [
    #     {
    #         "domain_name": "www.baidu.com",
    #         "product_name": "百度",
    #         "apply_name": "李彦宏",
    #         "apply_time": "2022-04-30",
    #         "itsm_id": "ID1001",
    #         "apply_reason": "测试1",
    #         "tech_name": "李彦宏",
    #         "ops_name": "李彦宏"
    #     },
    #     {
    #         "domain_name": "www.qq.com",
    #         "product_name": "腾讯",
    #         "apply_name": "马化腾",
    #         "apply_time": "2022-04-30",
    #         "itsm_id": "ID1002",
    #         "apply_reason": "测试1",
    #         "tech_name": "马化腾",
    #         "ops_name": "马化腾"
    #     }
    # ]
    #
    # ExcelExport.export_excel("文件名", columns_map, record_list)

    # 读取excel
    table = ExcelReader.read_excel("F:/workspace/~Python项目/pyscript/test\短信平台/token配置及使用情况统计", "Sheet1")
    rows = table.max_row  # 获取行数
    for i in range(1, rows):
        project_name = table.cell(row=i, column=1).value
        token = table.cell(row=i, column=2).value
        print(project_name, token)